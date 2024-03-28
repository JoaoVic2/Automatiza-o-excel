from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import pandas as pd
import os

# Obter a data atual e formatá-la no padrão YYYYMMDD
data_ontem = datetime.now() - timedelta(days=1)
data_ontem_apenas_data = data_ontem.date()
data_formatada = data_ontem_apenas_data.strftime("%Y%m%d") #datetime and timedelta 

nome_arquivo_csv = (f"FIDUCIAL_{data_formatada}.csv")  # Substitua pelo nome do seu arquivo CSV
dados = pd.read_csv(nome_arquivo_csv) #pd 

# Escrever os dados em um arquivo Excel
nome_arquivo_excel = f"FIDUCIAL_{data_formatada}.xlsx"  # Nome do arquivo Excel de saída
dados.to_excel(nome_arquivo_excel, index=False) #openpyxl

# Nome do arquivo da planilha existente
nome_arquivo = f"FIDUCIAL_{data_formatada}_separado.xlsx" #openpyxl

#nome do arquivo convertido 
nome_convertido = f"FIDUCIAL_{data_formatada}.csv" #openpyxl


dados_convertidos = pd.read_csv(nome_convertido, sep=';')#pd
nome_arquivo_excel = f"FIDUCIAL_{data_formatada}_separado.xlsx"
dados_convertidos.to_excel(nome_arquivo_excel, index=False)

# Carregar a planilha existente
workbook = load_workbook(f"FIDUCIAL_{data_formatada}_separado.xlsx") #openpyxl

# Selecionar a primeira planilha
sheet = workbook.active #openpyxl

# Definir o cabeçalho (títulos das colunas)
cabecalho = ['NOME DO ARQUIVO DE AUDIO','ID DO AGENTE','DATA E HORA DO CONTRATO','ID CHAMADA','FUNCIONARIO','EPS','ID CLIENTE','NOME DO SUPERVISOR','ID DO SUPERVISOR','FAIXA DE ATRASO','MOTIVO DO CONTATO - TABULACAO OPERACIONAL','PRODUTO','MOTIVO DE RECUSA','TIPO DE PESSOA','CIDADE','ESTADO','STATUS COBRANCA','OPERACAO','fila']
#openpyxl

# Inserir o cabeçalho na primeira linha da planilha
for index, titulo in enumerate(cabecalho, start=1):
    sheet[f"{get_column_letter(index)}1"] = titulo

# Salvar a planilha com o novo cabeçalho
workbook.save(nome_convertido)

os.remove(f"FIDUCIAL_{data_formatada}_separado.xlsx")
os.remove(f"FIDUCIAL_{data_formatada}.xlsx")